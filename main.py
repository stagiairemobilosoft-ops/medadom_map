import os
import re
import unicodedata
from difflib import SequenceMatcher

import openpyxl
import serpapi
from openpyxl.styles import Alignment


# 1. Excel
# ============================================================
FICHIER_ENTREE = "pharma.xlsx"
FICHIER_SORTIE = "pharma_resultats.xlsx"
SEUIL_SCORE_TROUVE = 0.50


# 2. Config SerpAPI
# ============================================================
api_key = os.getenv("SERPAPI_KEY")

if not api_key:
    raise RuntimeError("SERPAPI_KEY n'est pas définie")

client = serpapi.Client(
    api_key=api_key
)


# 3. Normalisation
# ============================================================
def normaliser(texte):

    if not texte:
        return ""

    texte = str(texte).lower()

    # Suppression des accents
    texte = unicodedata.normalize(
        "NFD",
        texte
    )

    texte = "".join(
        caractere
        for caractere in texte
        if unicodedata.category(caractere) != "Mn"
    )

    # Abréviations courantes
    texte = re.sub(r"\br\.\b", "rue", texte)
    texte = re.sub(r"\bav\.\b", "avenue", texte)
    texte = re.sub(r"\bbd\.\b", "boulevard", texte)

    # Ponctuation
    texte = re.sub(
        r"[^a-z0-9]",
        " ",
        texte
    )

    # Espaces multiples
    texte = re.sub(
        r"\s+",
        " ",
        texte
    )

    return texte.strip()


# 4. Requête
# ============================================================
def construire_query(
    location,
    adresse,
    code_postal,
    ville,
    pays
):

    elements = [
        location,
        adresse,
        code_postal,
        ville,
        pays
    ]

    elements_valides = []

    for element in elements:

        if element is not None:

            element = str(element).strip()

            if element:
                elements_valides.append(element)

    return " ".join(elements_valides)


# 5. Formatage des horaires
# ============================================================
def formater_horaires(horaires):

    if not horaires:
        return ""

    if isinstance(horaires, str):

        horaires = horaires.replace(
            ", ",
            "\n"
        )

        horaires = horaires.replace(
            ",",
            "\n"
        )

        return horaires.strip()

    if isinstance(horaires, list):

        return "\n".join(
            str(horaire).strip()
            for horaire in horaires
            if horaire
        )

    return str(horaires).strip()


# 6. Get Horaire
# ============================================================
def extraire_horaires(hours):

    horaires = {

        "lundi": "",
        "mardi": "",
        "mercredi": "",
        "jeudi": "",
        "vendredi": "",
        "samedi": "",
        "dimanche": ""
    }

    if not hours:
        return horaires

    correspondance_jours = {

        "monday": "lundi",
        "tuesday": "mardi",
        "wednesday": "mercredi",
        "thursday": "jeudi",
        "friday": "vendredi",
        "saturday": "samedi",
        "sunday": "dimanche",

        "lundi": "lundi",
        "mardi": "mardi",
        "mercredi": "mercredi",
        "jeudi": "jeudi",
        "vendredi": "vendredi",
        "samedi": "samedi",
        "dimanche": "dimanche"
    }

    for jour in hours:

        if not isinstance(jour, dict):
            continue

        for nom_jour, horaires_jour in jour.items():

            nom_jour_normalise = (
                str(nom_jour)
                .lower()
                .strip()
            )

            jour_francais = correspondance_jours.get(
                nom_jour_normalise
            )

            if jour_francais:

                horaires[jour_francais] = formater_horaires(
                    horaires_jour
                )

    return horaires


# 7. RÉCUPÉRATION DES DÉTAILS DE LA FICHE
# ============================================================
def recuperer_details_fiche(data_id):

    details = {

        "description": "",
        "site_web": "",
        "nombre_photos": 0,
        "nombre_photos_proprietaire": 0
    }

    if not data_id:

        print(
            "⚠️ Aucun data_id disponible"
        )

        return details


    # ========================================================
    # RECHERCHE DÉTAILLÉE DE LA FICHE
    # ========================================================

    try:

        print(
            "\n🔎 Recherche détaillée de la fiche..."
        )

        details_result = client.search({

            "engine": "google_maps",

            "type": "search",

            "data_id": data_id,

            "google_domain": "google.fr",

            "hl": "fr"
        })


        # ----------------------------------------------------
        # Affichage des clés disponibles
        # ----------------------------------------------------

        print(
            "Clés reçues :",
            list(details_result.keys())
        )


        # ----------------------------------------------------
        # Récupération de place_results
        # ----------------------------------------------------

        place_detail = details_result.get(
            "place_results",
            {}
        )


        if not isinstance(
            place_detail,
            dict
        ):

            place_detail = {}


        # ----------------------------------------------------
        # DESCRIPTION
        # ----------------------------------------------------

        description = place_detail.get(
            "description",
            ""
        )

        if not description:

            description = place_detail.get(
                "about",
                ""
            )

        if isinstance(
            description,
            dict
        ):

            description = (
                description.get(
                    "text",
                    ""
                )
            )


        details["description"] = (
            str(description).strip()
            if description
            else ""
        )


        # ----------------------------------------------------
        # SITE WEB
        # ----------------------------------------------------

        website = place_detail.get(
            "website",
            ""
        )

        if not website:

            website = place_detail.get(
                "website_url",
                ""
            )


        details["site_web"] = (
            str(website).strip()
            if website
            else ""
        )


        print(
            "Description trouvée :",
            bool(details["description"])
        )

        print(
            "Site web trouvé :",
            details["site_web"]
        )


    except Exception as e:

        print(
            "\n⚠️ ERREUR DÉTAILS FICHE"
        )

        print(e)


    # ========================================================
    # PHOTOS GOOGLE
    # ========================================================

    try:

        print(
            "\n📷 Recherche des photos Google..."
        )

        resultat_photos = client.search({

            "engine": "google_maps_photos",

            "data_id": data_id,

            "hl": "fr"
        })


        photos = resultat_photos.get(
            "photos",
            []
        )


        if isinstance(
            photos,
            list
        ):

            details["nombre_photos"] = len(
                photos
            )


        print(
            "Nombre de photos :",
            details["nombre_photos"]
        )


    except Exception as e:

        print(
            "⚠️ Erreur photos :"
        )

        print(e)


    # ========================================================
    # PHOTOS DU PROPRIÉTAIRE
    # ========================================================

    try:

        print(
            "\n👤 Recherche des photos propriétaire..."
        )

        resultat_proprietaire = client.search({

            "engine": "google_maps_photos",

            "data_id": data_id,

            "category_id": "CgIgARICEAE",

            "hl": "fr"
        })


        photos_proprietaire = (
            resultat_proprietaire.get(
                "photos",
                []
            )
        )


        if isinstance(
            photos_proprietaire,
            list
        ):

            details[
                "nombre_photos_proprietaire"
            ] = len(
                photos_proprietaire
            )


        print(
            "Nombre de photos propriétaire :",
            details[
                "nombre_photos_proprietaire"
            ]
        )


    except Exception as e:

        print(
            "⚠️ Erreur photos propriétaire :"
        )

        print(e)


    return details


# 8. RECHERCHE - LOCATION
# ============================================================
def rechercher_etablissement(
    location,
    adresse,
    code_postal,
    ville,
    pays
):

    query = construire_query(
        location,
        adresse,
        code_postal,
        ville,
        pays
    )

    print("\n")
    print("=" * 60)
    print("RECHERCHE")
    print("=" * 60)

    print(query)


    # Erreur de query
    # --------------------------------------------------------

    if not query:

        print(
            "🔴 Aucune donnée pour effectuer la recherche"
        )

        return {

            "recherche": "non trouvé",
            "score": 0.0,

            "adresse": "",
            "contact": "",

            "description": "",
            "site_web": "",

            "note_google": "",
            "nombre_avis_google": 0,

            "nombre_photos": 0,
            "nombre_photos_proprietaire": 0,

            "lundi": "",
            "mardi": "",
            "mercredi": "",
            "jeudi": "",
            "vendredi": "",
            "samedi": "",
            "dimanche": ""
        }


    # Recherche Google Maps
    # --------------------------------------------------------

    try:

        results = client.search({

            "engine": "google_maps",

            "type": "search",

            "q": query,

            "google_domain": "google.fr",

            "hl": "fr"
        })


    except Exception as e:

        print("\n🔴 ERREUR SERPAPI")

        print(e)

        return {

            "recherche": "non trouvé",
            "score": 0.0,

            "adresse": "",
            "contact": "",

            "description": "",
            "site_web": "",

            "note_google": "",
            "nombre_avis_google": 0,

            "nombre_photos": 0,
            "nombre_photos_proprietaire": 0,

            "lundi": "",
            "mardi": "",
            "mercredi": "",
            "jeudi": "",
            "vendredi": "",
            "samedi": "",
            "dimanche": ""
        }


    # Récupération de la fiche Google
    # --------------------------------------------------------

    place = results.get(
        "place_results",
        {}
    )


    if not place:

        print(
            "\n🔴 Aucune fiche Google trouvée"
        )

        return {

            "recherche": "non trouvé",
            "score": 0.0,

            "adresse": "",
            "contact": "",

            "description": "",
            "site_web": "",

            "note_google": "",
            "nombre_avis_google": 0,

            "nombre_photos": 0,
            "nombre_photos_proprietaire": 0,

            "lundi": "",
            "mardi": "",
            "mercredi": "",
            "jeudi": "",
            "vendredi": "",
            "samedi": "",
            "dimanche": ""
        }


    # Données Google
    # --------------------------------------------------------

    nom_google = place.get(
        "title",
        ""
    )

    adresse_google = place.get(
        "address",
        ""
    )

    telephone_google = place.get(
        "phone",
        ""
    )

    horaires_google = place.get(
        "hours",
        []
    )


    # ========================================================
    # DATA ID
    # ========================================================

    data_id = place.get(
        "data_id",
        ""
    )


    print(
        "\nDATA ID :",
        data_id
    )


    # ========================================================
    # DESCRIPTION + SITE WEB + PHOTOS
    # ========================================================

    details = recuperer_details_fiche(
        data_id
    )


    description_google = details[
        "description"
    ]

    site_web_google = details[
        "site_web"
    ]

    nombre_photos = details[
        "nombre_photos"
    ]

    nombre_photos_proprietaire = details[
        "nombre_photos_proprietaire"
    ]


    # ========================================================
    # FALLBACK DESCRIPTION
    # ========================================================

    if not description_google:

        description_google = place.get(
            "description",
            ""
        )


    # ========================================================
    # FALLBACK SITE WEB
    # ========================================================

    if not site_web_google:

        site_web_google = place.get(
            "website",
            ""
        )


    # ========================================================
    # NOTE GOOGLE + NOMBRE D'AVIS
    # ========================================================

    note_google = place.get(
        "rating",
        ""
    )

    nombre_avis_google = place.get(
        "reviews",
        0
    )


    print("\nFICHE GOOGLE")
    print("-" * 60)

    print(
        "Nom       :",
        nom_google
    )

    print(
        "Adresse   :",
        adresse_google
    )

    print(
        "Téléphone :",
        telephone_google
    )

    print(
        "Description :",
        description_google
    )

    print(
        "Site web  :",
        site_web_google
    )

    print(
        "Note Google :",
        note_google
    )

    print(
        "Nombre avis :",
        nombre_avis_google
    )

    print(
        "Nombre photos :",
        nombre_photos
    )

    print(
        "Nombre photos propriétaire :",
        nombre_photos_proprietaire
    )


    # SCORE RECHERCHE
    # COMPARAISON DU NOM
    # ========================================================

    nom_source = normaliser(
        location
    )

    nom_google_normalise = normaliser(
        nom_google
    )


    if nom_source and nom_google_normalise:

        similarite_nom = SequenceMatcher(
            None,
            nom_source,
            nom_google_normalise
        ).ratio()

    else:

        similarite_nom = 0.0


    # ========================================================
    # COMPARAISON DE L'ADRESSE
    # ========================================================

    adresse_source = normaliser(
        adresse
    )

    adresse_google_normalisee = normaliser(
        adresse_google
    )


    if adresse_source and adresse_google_normalisee:

        similarite_adresse = SequenceMatcher(
            None,
            adresse_source,
            adresse_google_normalisee
        ).ratio()

    else:

        similarite_adresse = 1.0


    # ========================================================
    # CODE POSTAL
    # ========================================================

    if code_postal:

        code_postal_normalise = normaliser(
            code_postal
        )

        code_postal_trouve = (
            code_postal_normalise
            in adresse_google_normalisee
        )

        score_code_postal = (
            1.0
            if code_postal_trouve
            else 0.0
        )

    else:

        score_code_postal = 1.0


    # ========================================================
    # VILLE
    # ========================================================

    if ville:

        ville_normalisee = normaliser(
            ville
        )

        ville_trouvee = (
            ville_normalisee
            in adresse_google_normalisee
        )

        score_ville = (
            1.0
            if ville_trouvee
            else 0.0
        )

    else:

        score_ville = 1.0


    # ========================================================
    # SCORE GLOBAL
    # ========================================================

    score_global = (

        similarite_nom * 0.50

        + similarite_adresse * 0.30

        + score_ville * 0.10

        + score_code_postal * 0.10
    )


    # ========================================================
    # AFFICHAGE DES SCORES
    # ========================================================

    print("\nCOMPARAISON")
    print("=" * 60)

    print(
        f"Similarité nom       : "
        f"{similarite_nom:.2f}"
    )

    print(
        f"Similarité adresse   : "
        f"{similarite_adresse:.2f}"
    )

    print(
        f"Correspondance ville : "
        f"{score_ville:.2f}"
    )

    print(
        f"Code postal          : "
        f"{score_code_postal:.2f}"
    )

    print("-" * 60)

    print(
        f"SCORE GLOBAL         : "
        f"{score_global:.2f}"
    )


    # ========================================================
    # DÉTERMINATION DU STATUT
    # ========================================================

    if score_global > SEUIL_SCORE_TROUVE:

        statut = "trouvé"

    elif score_global > 0:

        statut = "probable"

    else:

        statut = "non trouvé"


    print(
        f"RÉSULTAT             : {statut}"
    )


    # ========================================================
    # HORAIRES
    # ========================================================

    horaires = extraire_horaires(
        horaires_google
    )


    # ========================================================
    # RETOUR
    # ========================================================

    return {

        "recherche": statut,

        "score": score_global,

        "adresse": adresse_google,

        "contact": telephone_google,

        "description": description_google,

        "site_web": site_web_google,

        "note_google": note_google,

        "nombre_avis_google": nombre_avis_google,

        "nombre_photos": nombre_photos,

        "nombre_photos_proprietaire":
            nombre_photos_proprietaire,

        "lundi": horaires["lundi"],
        "mardi": horaires["mardi"],
        "mercredi": horaires["mercredi"],
        "jeudi": horaires["jeudi"],
        "vendredi": horaires["vendredi"],
        "samedi": horaires["samedi"],
        "dimanche": horaires["dimanche"]
    }


# 9. Lecture du fichier Excel
# ============================================================

print("\n")
print("=" * 60)
print("LECTURE DU FICHIER")
print("=" * 60)

print(
    "Fichier :",
    FICHIER_ENTREE
)

classeur = openpyxl.load_workbook(
    FICHIER_ENTREE
)

feuille = classeur.active


# 10. Colonne de query
# ============================================================

headers = {}

for cellule in feuille[1]:

    if cellule.value is not None:

        nom_colonne = (
            str(cellule.value)
            .strip()
            .lower()
        )

        headers[nom_colonne] = cellule.column


colonnes_obligatoires = [

    "location",
    "adress",
    "code_postal",
    "ville",
    "pays"
]


for colonne in colonnes_obligatoires:

    if colonne not in headers:

        raise RuntimeError(
            f"Colonne obligatoire absente : "
            f"{colonne}"
        )


# 11. Génération des résultats
# ============================================================

classeur_sortie = openpyxl.Workbook()

feuille_sortie = (
    classeur_sortie.active
)

feuille_sortie.title = "Résultats"


# 12. Colonnes du fichier de résultats
# ============================================================

colonnes_sortie = [

    "location",
    "adress",
    "code_postal",
    "ville",
    "pays",

    "recherche",

    "adresse",
    "contact",

    "description",
    "site_web",

    "note_google",
    "nombre_avis_google",

    "nombre_photos",
    "nombre_photos_proprietaire",

    "lundi",
    "mardi",
    "mercredi",
    "jeudi",
    "vendredi",
    "samedi",
    "dimanche",

    "score"
]


feuille_sortie.append(
    colonnes_sortie
)


# 13. Formes des entêtes
# ============================================================

for cellule in feuille_sortie[1]:

    cellule.alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True
    )


# 14. Colonne score automatique
# ============================================================

COLONNE_SCORE = (
    colonnes_sortie.index("score")
    + 1
)


# 15. Traitement des lignes
# ============================================================

nombre_total = 0
nombre_trouve = 0
nombre_probable = 0
nombre_non_trouve = 0


print("\n")
print("=" * 60)
print("DÉBUT DU TRAITEMENT")
print("=" * 60)


for numero_ligne in range(
    2,
    feuille.max_row + 1
):

    nombre_total += 1

    print("\n")
    print(
        "#" * 20,
        f"LIGNE {numero_ligne}",
        "#" * 20
    )


    # --------------------------------------------------------
    # Lecture des données
    # --------------------------------------------------------

    location = feuille.cell(
        numero_ligne,
        headers["location"]
    ).value

    adresse = feuille.cell(
        numero_ligne,
        headers["adress"]
    ).value

    code_postal = feuille.cell(
        numero_ligne,
        headers["code_postal"]
    ).value

    ville = feuille.cell(
        numero_ligne,
        headers["ville"]
    ).value

    pays = feuille.cell(
        numero_ligne,
        headers["pays"]
    ).value


    # --------------------------------------------------------
    # Nettoyage
    # --------------------------------------------------------

    location = (
        str(location).strip()
        if location is not None
        else ""
    )

    adresse = (
        str(adresse).strip()
        if adresse is not None
        else ""
    )

    code_postal = (
        str(code_postal).strip()
        if code_postal is not None
        else ""
    )

    ville = (
        str(ville).strip()
        if ville is not None
        else ""
    )

    pays = (
        str(pays).strip()
        if pays is not None
        else ""
    )


    # --------------------------------------------------------
    # Ligne vide
    # --------------------------------------------------------

    if not any([
        location,
        adresse,
        code_postal,
        ville,
        pays
    ]):

        print(
            f"Ligne {numero_ligne} vide"
        )

        feuille_sortie.append([

            location,
            adresse,
            code_postal,
            ville,
            pays,

            "non trouvé",

            "",
            "",

            "",
            "",

            "",
            "",

            0,
            0,

            "",
            "",
            "",
            "",
            "",
            "",
            "",

            0.0
        ])

        nombre_non_trouve += 1

        continue


    # --------------------------------------------------------
    # Recherche Google
    # --------------------------------------------------------

    resultat = rechercher_etablissement(

        location,

        adresse,

        code_postal,

        ville,

        pays
    )


    # --------------------------------------------------------
    # Ajouter toujours la ligne
    # --------------------------------------------------------

    if resultat is None:

        resultat = {

            "recherche": "non trouvé",
            "score": 0.0,

            "adresse": "",
            "contact": "",

            "description": "",
            "site_web": "",

            "note_google": "",
            "nombre_avis_google": 0,

            "nombre_photos": 0,
            "nombre_photos_proprietaire": 0,

            "lundi": "",
            "mardi": "",
            "mercredi": "",
            "jeudi": "",
            "vendredi": "",
            "samedi": "",
            "dimanche": ""
        }


    # --------------------------------------------------------
    # Comptage
    # --------------------------------------------------------

    if resultat["recherche"] == "trouvé":

        nombre_trouve += 1

    elif resultat["recherche"] == "probable":

        nombre_probable += 1

    else:

        nombre_non_trouve += 1


    # --------------------------------------------------------
    # Ajout dans Excel
    # --------------------------------------------------------

    feuille_sortie.append([

        # Données originales
        location,
        adresse,
        code_postal,
        ville,
        pays,

        # Résultat
        resultat["recherche"],

        # Google
        resultat["adresse"],
        resultat["contact"],

        # Description
        resultat["description"],

        # Site web
        resultat["site_web"],

        # Avis Google
        resultat["note_google"],
        resultat["nombre_avis_google"],

        # Photos
        resultat["nombre_photos"],
        resultat["nombre_photos_proprietaire"],

        # Horaires
        resultat["lundi"],
        resultat["mardi"],
        resultat["mercredi"],
        resultat["jeudi"],
        resultat["vendredi"],
        resultat["samedi"],
        resultat["dimanche"],

        # Score
        resultat["score"]
    ])


    # --------------------------------------------------------
    # Alignement et retour à la ligne
    # --------------------------------------------------------

    derniere_ligne = (
        feuille_sortie.max_row
    )

    for cellule in feuille_sortie[
        derniere_ligne
    ]:

        cellule.alignment = Alignment(
            vertical="top",
            wrap_text=True
        )


    # --------------------------------------------------------
    # Score sous forme de %
    # --------------------------------------------------------

    feuille_sortie.cell(
        derniere_ligne,
        COLONNE_SCORE
    ).number_format = "0%"


# 16. Largeurs des cellules
# ============================================================

largeurs = {

    "A": 30,
    "B": 35,
    "C": 15,
    "D": 30,
    "E": 15,

    "F": 18,

    "G": 40,
    "H": 20,

    "I": 50,
    "J": 40,

    "K": 15,
    "L": 20,

    "M": 20,
    "N": 28,

    "O": 22,
    "P": 22,
    "Q": 22,
    "R": 22,
    "S": 22,
    "T": 22,
    "U": 22,

    "V": 12
}


for colonne, largeur in largeurs.items():

    feuille_sortie.column_dimensions[
        colonne
    ].width = largeur


# 17. Hauteurs des lignes
# ============================================================

for numero_ligne in range(
    2,
    feuille_sortie.max_row + 1
):

    feuille_sortie.row_dimensions[
        numero_ligne
    ].height = 40


# 18. Sauvegarde
# ============================================================

classeur_sortie.save(
    FICHIER_SORTIE
)


# 19. Résumé
# ============================================================

print("\n")
print("=" * 60)
print("TRAITEMENT TERMINÉ")
print("=" * 60)

print(
    f"Nombre de lignes analysées : "
    f"{nombre_total}"
)

print(
    f"🟢 Trouvé (> 50 %)          : "
    f"{nombre_trouve}"
)

print(
    f"🟡 Probable (0–50 %)        : "
    f"{nombre_probable}"
)

print(
    f"🔴 Non trouvé (0 %)         : "
    f"{nombre_non_trouve}"
)

print(
    f"\nFichier créé : "
    f"{FICHIER_SORTIE}"
)

print("=" * 60)
