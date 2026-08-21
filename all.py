import os
import re
import unicodedata
from datetime import datetime
from difflib import SequenceMatcher

import openpyxl
import serpapi
from openpyxl.styles import Alignment


# ============================================================
# 1. CONFIGURATION EXCEL
# ============================================================

FICHIER_ENTREE = "pharma.xlsx"
FICHIER_SORTIE = "pharma_resultats.xlsx"

SEUIL_SCORE_TROUVE = 0.50


# ============================================================
# 2. CONFIGURATION SERPAPI
# ============================================================

api_key = os.getenv("SERPAPI_KEY")

if not api_key:
    raise RuntimeError("SERPAPI_KEY n'est pas définie")

client = serpapi.Client(
    api_key=api_key
)


# ============================================================
# 3. NORMALISATION
# ============================================================

def normaliser(texte):

    if not texte:
        return ""

    texte = str(texte).lower()

    # Suppression des accents
    texte = unicodedata.normalize(
        "NFD",
        texte
    )

    texte = "".join(
        caractere
        for caractere in texte
        if unicodedata.category(caractere) != "Mn"
    )

    # Abréviations courantes
    texte = re.sub(
        r"\br\.\b",
        "rue",
        texte
    )

    texte = re.sub(
        r"\bav\.\b",
        "avenue",
        texte
    )

    texte = re.sub(
        r"\bbd\.\b",
        "boulevard",
        texte
    )

    # Ponctuation
    texte = re.sub(
        r"[^a-z0-9]",
        " ",
        texte
    )

    # Espaces multiples
    texte = re.sub(
        r"\s+",
        " ",
        texte
    )

    return texte.strip()


# ============================================================
# 4. CONSTRUCTION DE LA REQUÊTE
# ============================================================

def construire_query(
    location,
    adresse,
    code_postal,
    ville,
    pays
):

    elements = [
        location,
        adresse,
        code_postal,
        ville,
        pays
    ]

    elements_valides = []

    for element in elements:

        if element is not None:

            element = str(element).strip()

            if element:
                elements_valides.append(element)

    return " ".join(elements_valides)


# ============================================================
# 5. FORMATAGE DES HORAIRES
# ============================================================

def formater_horaires(horaires):

    if not horaires:
        return ""

    if isinstance(horaires, str):

        horaires = horaires.replace(
            ", ",
            "\n"
        )

        horaires = horaires.replace(
            ",",
            "\n"
        )

        return horaires.strip()

    if isinstance(horaires, list):

        return "\n".join(
            str(horaire).strip()
            for horaire in horaires
            if horaire
        )

    return str(horaires).strip()


# ============================================================
# 6. EXTRACTION DES HORAIRES
# ============================================================

def extraire_horaires(hours):

    horaires = {

        "lundi": "",
        "mardi": "",
        "mercredi": "",
        "jeudi": "",
        "vendredi": "",
        "samedi": "",
        "dimanche": ""
    }

    if not hours:
        return horaires

    correspondance_jours = {

        "monday": "lundi",
        "tuesday": "mardi",
        "wednesday": "mercredi",
        "thursday": "jeudi",
        "friday": "vendredi",
        "saturday": "samedi",
        "sunday": "dimanche",

        "lundi": "lundi",
        "mardi": "mardi",
        "mercredi": "mercredi",
        "jeudi": "jeudi",
        "vendredi": "vendredi",
        "samedi": "samedi",
        "dimanche": "dimanche"
    }

    for jour in hours:

        if not isinstance(jour, dict):
            continue

        for nom_jour, horaires_jour in jour.items():

            nom_jour_normalise = (
                str(nom_jour)
                .lower()
                .strip()
            )

            jour_francais = correspondance_jours.get(
                nom_jour_normalise
            )

            if jour_francais:

                horaires[jour_francais] = (
                    formater_horaires(
                        horaires_jour
                    )
                )

    return horaires


# ============================================================
# 7. CONVERSION DATE GOOGLE EN FRANÇAIS
# ============================================================

def convertir_date_google(date_value):

    if not date_value:
        return ""

    date_value = str(
        date_value
    ).strip()

    # ========================================================
    # CAS ISO
    #
    # Exemple :
    # 2026-08-18T21:22:18Z
    # ========================================================

    try:

        date_obj = datetime.fromisoformat(
            date_value.replace(
                "Z",
                "+00:00"
            )
        )

        mois_francais = {

            1: "janvier",
            2: "février",
            3: "mars",
            4: "avril",
            5: "mai",
            6: "juin",
            7: "juillet",
            8: "août",
            9: "septembre",
            10: "octobre",
            11: "novembre",
            12: "décembre"
        }

        return (
            f"{date_obj.day} "
            f"{mois_francais[date_obj.month]} "
            f"{date_obj.year}"
        )

    except Exception:
        pass

    # ========================================================
    # CAS :
    # Aug 12, 2026
    # ========================================================

    mois_anglais = {

        "Jan": "janvier",
        "Feb": "février",
        "Mar": "mars",
        "Apr": "avril",
        "May": "mai",
        "Jun": "juin",
        "Jul": "juillet",
        "Aug": "août",
        "Sep": "septembre",
        "Oct": "octobre",
        "Nov": "novembre",
        "Dec": "décembre"
    }

    correspondance = re.match(
        r"([A-Za-z]{3})\s+(\d{1,2}),\s+(\d{4})",
        date_value
    )

    if correspondance:

        mois = correspondance.group(1)
        jour = correspondance.group(2)
        annee = correspondance.group(3)

        mois_fr = mois_anglais.get(
            mois,
            mois
        )

        return (
            f"{jour} "
            f"{mois_fr} "
            f"{annee}"
        )

    # ========================================================
    # CAS :
    # Aug 12 2026
    # ========================================================

    correspondance = re.match(
        r"([A-Za-z]{3})\s+(\d{1,2})\s+(\d{4})",
        date_value
    )

    if correspondance:

        mois = correspondance.group(1)
        jour = correspondance.group(2)
        annee = correspondance.group(3)

        mois_fr = mois_anglais.get(
            mois,
            mois
        )

        return (
            f"{jour} "
            f"{mois_fr} "
            f"{annee}"
        )

    return date_value


# ============================================================
# 8. RECHERCHE PHOTOS GOOGLE
# ============================================================

def rechercher_photos(data_id):

    nombre_photos = 0
    nombre_photos_proprietaire = 0

    if not data_id:

        return (
            nombre_photos,
            nombre_photos_proprietaire
        )

    # ========================================================
    # TOUTES LES PHOTOS
    # ========================================================

    try:

        print(
            "\n📷 Recherche des photos Google..."
        )

        resultat_photos = client.search({

            "engine": "google_maps_photos",

            "data_id": data_id,

            "hl": "fr"
        })

        photos = resultat_photos.get(
            "photos",
            []
        )

        if isinstance(
            photos,
            list
        ):

            nombre_photos = len(
                photos
            )

        print(
            "Nombre de photos récupérées :",
            nombre_photos
        )

    except Exception as e:

        print(
            "⚠️ Erreur récupération photos :"
        )

        print(e)

    # ========================================================
    # PHOTOS DU PROPRIÉTAIRE
    # ========================================================

    try:

        print(
            "👤 Recherche des photos propriétaire..."
        )

        resultat_proprietaire = client.search({

            "engine": "google_maps_photos",

            "data_id": data_id,

            "category_id": "CgIgARICEAE",

            "hl": "fr"
        })

        photos_proprietaire = (
            resultat_proprietaire.get(
                "photos",
                []
            )
        )

        if isinstance(
            photos_proprietaire,
            list
        ):

            nombre_photos_proprietaire = len(
                photos_proprietaire
            )

        print(
            "Nombre de photos propriétaire :",
            nombre_photos_proprietaire
        )

    except Exception as e:

        print(
            "⚠️ Erreur récupération photos propriétaire :"
        )

        print(e)

    return (
        nombre_photos,
        nombre_photos_proprietaire
    )


# ============================================================
# 9. RECHERCHE DATE DU DERNIER POST GOOGLE
# ============================================================

def rechercher_date_dernier_post(data_id):

    if not data_id:

        return ""

    try:

        print(
            "\n📢 Recherche du dernier post Google..."
        )

        resultat_posts = client.search({

            "engine": "google_maps_posts",

            "data_id": data_id,

            "hl": "en"
        })

        posts = resultat_posts.get(
            "posts",
            []
        )

        if not posts:

            print(
                "Aucun post Google trouvé"
            )

            return ""

        # Le premier post est le plus récent
        dernier = posts[0]

        print(
            "\nDONNÉES DU DERNIER POST"
        )

        print(
            "-" * 60
        )

        print(
            "time :",
            dernier.get(
                "time",
                ""
            )
        )

        print(
            "from :",
            dernier.get(
                "from",
                ""
            )
        )

        print(
            "to :",
            dernier.get(
                "to",
                ""
            )
        )

        print(
            "posted_at :",
            dernier.get(
                "posted_at",
                ""
            )
        )

        print(
            "posted_at_text :",
            dernier.get(
                "posted_at_text",
                ""
            )
        )

        # ====================================================
        # DATE DU POST
        # ====================================================

        posted_at = dernier.get(
            "posted_at",
            ""
        )

        if posted_at:

            date_convertie = (
                convertir_date_google(
                    posted_at
                )
            )

            print(
                "Date du dernier post :",
                date_convertie
            )

            return date_convertie

        # ====================================================
        # FALLBACK posted_at_text
        # ====================================================

        posted_at_text = dernier.get(
            "posted_at_text",
            ""
        )

        if posted_at_text:

            # Si Google donne directement une date
            # on la conserve.
            if not re.search(
                r"\b(?:ago|day|days|hour|hours|week|weeks|month|months|year|years)\b",
                posted_at_text,
                re.IGNORECASE
            ):

                date_convertie = (
                    convertir_date_google(
                        posted_at_text
                    )
                )

                print(
                    "Date du dernier post :",
                    date_convertie
                )

                return date_convertie

        print(
            "Date du dernier post introuvable"
        )

        return ""

    except Exception as e:

        print(
            "⚠️ Erreur récupération date dernier post :"
        )

        print(e)

        return ""


# ============================================================
# 10. RECHERCHE ÉTABLISSEMENT
# ============================================================

def rechercher_etablissement(
    location,
    adresse,
    code_postal,
    ville,
    pays
):

    query = construire_query(

        location,

        adresse,

        code_postal,

        ville,

        pays
    )

    print("\n")

    print(
        "=" * 60
    )

    print(
        "RECHERCHE"
    )

    print(
        "=" * 60
    )

    print(query)

    # ========================================================
    # QUERY VIDE
    # ========================================================

    if not query:

        print(
            "🔴 Aucune donnée pour effectuer la recherche"
        )

        return {

            "recherche": "non trouvé",

            "score": 0.0,

            "adresse": "",

            "contact": "",

            "site_web": "",

            "date_dernier_post": "",

            "note_google": "",

            "nombre_avis_google": 0,

            "nombre_photos": 0,

            "nombre_photos_proprietaire": 0,

            "lundi": "",
            "mardi": "",
            "mercredi": "",
            "jeudi": "",
            "vendredi": "",
            "samedi": "",
            "dimanche": ""
        }

    # ========================================================
    # RECHERCHE GOOGLE MAPS
    # ========================================================

    try:

        results = client.search({

            "engine": "google_maps",

            "type": "search",

            "q": query,

            "google_domain": "google.fr",

            "hl": "fr"
        })

    except Exception as e:

        print(
            "\n🔴 ERREUR SERPAPI"
        )

        print(e)

        return {

            "recherche": "non trouvé",

            "score": 0.0,

            "adresse": "",

            "contact": "",

            "site_web": "",

            "date_dernier_post": "",

            "note_google": "",

            "nombre_avis_google": 0,

            "nombre_photos": 0,

            "nombre_photos_proprietaire": 0,

            "lundi": "",
            "mardi": "",
            "mercredi": "",
            "jeudi": "",
            "vendredi": "",
            "samedi": "",
            "dimanche": ""
        }

    # ========================================================
    # RÉCUPÉRATION FICHE GOOGLE
    # ========================================================

    place = results.get(
        "place_results",
        {}
    )

    if not place:

        print(
            "\n🔴 Aucune fiche Google trouvée"
        )

        return {

            "recherche": "non trouvé",

            "score": 0.0,

            "adresse": "",

            "contact": "",

            "site_web": "",

            "date_dernier_post": "",

            "note_google": "",

            "nombre_avis_google": 0,

            "nombre_photos": 0,

            "nombre_photos_proprietaire": 0,

            "lundi": "",
            "mardi": "",
            "mercredi": "",
            "jeudi": "",
            "vendredi": "",
            "samedi": "",
            "dimanche": ""
        }

    # ========================================================
    # DONNÉES GOOGLE
    # ========================================================

    nom_google = place.get(
        "title",
        ""
    )

    adresse_google = place.get(
        "address",
        ""
    )

    telephone_google = place.get(
        "phone",
        ""
    )

    horaires_google = place.get(
        "hours",
        []
    )

    # ========================================================
    # SITE WEB
    # ========================================================

    site_web_google = place.get(
        "website",
        ""
    )

    # ========================================================
    # NOTE GOOGLE
    # ========================================================

    note_google = place.get(
        "rating",
        ""
    )

    nombre_avis_google = place.get(
        "reviews",
        0
    )

    # ========================================================
    # DATA ID
    # ========================================================

    data_id = place.get(
        "data_id",
        ""
    )

    # ========================================================
    # DATE DU DERNIER POST
    # ========================================================

    date_dernier_post = (
        rechercher_date_dernier_post(
            data_id
        )
    )

    # ========================================================
    # PHOTOS
    # ========================================================

    (
        nombre_photos,
        nombre_photos_proprietaire
    ) = rechercher_photos(
        data_id
    )

    # ========================================================
    # AFFICHAGE
    # ========================================================

    print(
        "\nFICHE GOOGLE"
    )

    print(
        "-" * 60
    )

    print(
        "Nom       :",
        nom_google
    )

    print(
        "Adresse   :",
        adresse_google
    )

    print(
        "Téléphone :",
        telephone_google
    )

    print(
        "Site web  :",
        site_web_google
    )

    print(
        "Date dernier post :",
        date_dernier_post
    )

    print(
        "Note Google :",
        note_google
    )

    print(
        "Nombre avis :",
        nombre_avis_google
    )

    print(
        "Nombre photos :",
        nombre_photos
    )

    print(
        "Photos propriétaire :",
        nombre_photos_proprietaire
    )

    # ========================================================
    # COMPARAISON NOM
    # ========================================================

    nom_source = normaliser(
        location
    )

    nom_google_normalise = normaliser(
        nom_google
    )

    if (
        nom_source
        and nom_google_normalise
    ):

        similarite_nom = SequenceMatcher(

            None,

            nom_source,

            nom_google_normalise

        ).ratio()

    else:

        similarite_nom = 0.0

    # ========================================================
    # COMPARAISON ADRESSE
    # ========================================================

    adresse_source = normaliser(
        adresse
    )

    adresse_google_normalisee = normaliser(
        adresse_google
    )

    if (
        adresse_source
        and adresse_google_normalisee
    ):

        similarite_adresse = SequenceMatcher(

            None,

            adresse_source,

            adresse_google_normalisee

        ).ratio()

    else:

        similarite_adresse = 1.0

    # ========================================================
    # CODE POSTAL
    # ========================================================

    if code_postal:

        code_postal_normalise = normaliser(
            code_postal
        )

        code_postal_trouve = (

            code_postal_normalise

            in adresse_google_normalisee
        )

        score_code_postal = (

            1.0

            if code_postal_trouve

            else 0.0
        )

    else:

        score_code_postal = 1.0

    # ========================================================
    # VILLE
    # ========================================================

    if ville:

        ville_normalisee = normaliser(
            ville
        )

        ville_trouvee = (

            ville_normalisee

            in adresse_google_normalisee
        )

        score_ville = (

            1.0

            if ville_trouvee

            else 0.0
        )

    else:

        score_ville = 1.0

    # ========================================================
    # SCORE GLOBAL
    # ========================================================

    score_global = (

        similarite_nom * 0.50

        + similarite_adresse * 0.30

        + score_ville * 0.10

        + score_code_postal * 0.10
    )

    # ========================================================
    # AFFICHAGE SCORES
    # ========================================================

    print(
        "\nCOMPARAISON"
    )

    print(
        "=" * 60
    )

    print(
        f"Similarité nom       : "
        f"{similarite_nom:.2f}"
    )

    print(
        f"Similarité adresse   : "
        f"{similarite_adresse:.2f}"
    )

    print(
        f"Correspondance ville : "
        f"{score_ville:.2f}"
    )

    print(
        f"Code postal          : "
        f"{score_code_postal:.2f}"
    )

    print(
        "-" * 60
    )

    print(
        f"SCORE GLOBAL         : "
        f"{score_global:.2f}"
    )

    # ========================================================
    # STATUT
    # ========================================================

    if score_global > SEUIL_SCORE_TROUVE:

        statut = "trouvé"

    elif score_global > 0:

        statut = "probable"

    else:

        statut = "non trouvé"

    print(
        f"RÉSULTAT             : {statut}"
    )

    # ========================================================
    # HORAIRES
    # ========================================================

    horaires = extraire_horaires(
        horaires_google
    )

    # ========================================================
    # RETOUR
    # ========================================================

    return {

        "recherche": statut,

        "score": score_global,

        "adresse": adresse_google,

        "contact": telephone_google,

        "site_web": site_web_google,

        "date_dernier_post":
            date_dernier_post,

        "note_google": note_google,

        "nombre_avis_google":
            nombre_avis_google,

        "nombre_photos":
            nombre_photos,

        "nombre_photos_proprietaire":
            nombre_photos_proprietaire,

        "lundi":
            horaires["lundi"],

        "mardi":
            horaires["mardi"],

        "mercredi":
            horaires["mercredi"],

        "jeudi":
            horaires["jeudi"],

        "vendredi":
            horaires["vendredi"],

        "samedi":
            horaires["samedi"],

        "dimanche":
            horaires["dimanche"]
    }


# ============================================================
# 11. LECTURE DU FICHIER EXCEL
# ============================================================

print("\n")

print(
    "=" * 60
)

print(
    "LECTURE DU FICHIER"
)

print(
    "=" * 60
)

print(
    "Fichier :",
    FICHIER_ENTREE
)

classeur = openpyxl.load_workbook(
    FICHIER_ENTREE
)

feuille = classeur.active


# ============================================================
# 12. LECTURE DES COLONNES
# ============================================================

headers = {}

for cellule in feuille[1]:

    if cellule.value is not None:

        nom_colonne = (
            str(cellule.value)
            .strip()
            .lower()
        )

        headers[nom_colonne] = (
            cellule.column
        )


colonnes_obligatoires = [

    "location",

    "adress",

    "code_postal",

    "ville",

    "pays"
]


for colonne in colonnes_obligatoires:

    if colonne not in headers:

        raise RuntimeError(

            f"Colonne obligatoire absente : "
            f"{colonne}"
        )


# ============================================================
# 13. CRÉATION DU FICHIER DE SORTIE
# ============================================================

classeur_sortie = (
    openpyxl.Workbook()
)

feuille_sortie = (
    classeur_sortie.active
)

feuille_sortie.title = "Résultats"


# ============================================================
# 14. COLONNES DE SORTIE
# ============================================================

colonnes_sortie = [

    "location",
    "adress",
    "code_postal",
    "ville",
    "pays",

    "recherche",

    "adresse",
    "contact",

    "site_web",

    "date_dernier_post",

    "note_google",
    "nombre_avis_google",

    "nombre_photos",
    "nombre_photos_proprietaire",

    "lundi",
    "mardi",
    "mercredi",
    "jeudi",
    "vendredi",
    "samedi",
    "dimanche",

    "score"
]


feuille_sortie.append(
    colonnes_sortie
)


# ============================================================
# 15. STYLE DES ENTÊTES
# ============================================================

for cellule in feuille_sortie[1]:

    cellule.alignment = Alignment(

        horizontal="center",

        vertical="center",

        wrap_text=True
    )


# ============================================================
# 16. COLONNE SCORE
# ============================================================

COLONNE_SCORE = (

    colonnes_sortie.index(
        "score"
    ) + 1
)


# ============================================================
# 17. COMPTEURS
# ============================================================

nombre_total = 0
nombre_trouve = 0
nombre_probable = 0
nombre_non_trouve = 0


print("\n")

print(
    "=" * 60
)

print(
    "DÉBUT DU TRAITEMENT"
)

print(
    "=" * 60
)


# ============================================================
# 18. TRAITEMENT DES LIGNES
# ============================================================

for numero_ligne in range(

    2,

    feuille.max_row + 1

):

    nombre_total += 1

    print("\n")

    print(

        "#" * 20,

        f"LIGNE {numero_ligne}",

        "#" * 20
    )

    # ========================================================
    # LECTURE DES DONNÉES
    # ========================================================

    location = feuille.cell(

        numero_ligne,

        headers["location"]

    ).value

    adresse = feuille.cell(

        numero_ligne,

        headers["adress"]

    ).value

    code_postal = feuille.cell(

        numero_ligne,

        headers["code_postal"]

    ).value

    ville = feuille.cell(

        numero_ligne,

        headers["ville"]

    ).value

    pays = feuille.cell(

        numero_ligne,

        headers["pays"]

    ).value

    # ========================================================
    # NETTOYAGE
    # ========================================================

    location = (

        str(location).strip()

        if location is not None

        else ""
    )

    adresse = (

        str(adresse).strip()

        if adresse is not None

        else ""
    )

    code_postal = (

        str(code_postal).strip()

        if code_postal is not None

        else ""
    )

    ville = (

        str(ville).strip()

        if ville is not None

        else ""
    )

    pays = (

        str(pays).strip()

        if pays is not None

        else ""
    )

    # ========================================================
    # LIGNE VIDE
    # ========================================================

    if not any([

        location,

        adresse,

        code_postal,

        ville,

        pays

    ]):

        print(
            f"Ligne {numero_ligne} vide"
        )

        feuille_sortie.append([

            location,

            adresse,

            code_postal,

            ville,

            pays,

            "non trouvé",

            "",

            "",

            "",

            "",

            "",

            0,

            0,

            0,

            "",

            "",

            "",

            "",

            "",

            "",

            "",

            "",

            0.0
        ])

        nombre_non_trouve += 1

        continue

    # ========================================================
    # RECHERCHE GOOGLE
    # ========================================================

    resultat = rechercher_etablissement(

        location,

        adresse,

        code_postal,

        ville,

        pays
    )

    # ========================================================
    # VALEURS PAR DÉFAUT
    # ========================================================

    if resultat is None:

        resultat = {

            "recherche": "non trouvé",

            "score": 0.0,

            "adresse": "",

            "contact": "",

            "site_web": "",

            "date_dernier_post": "",

            "note_google": "",

            "nombre_avis_google": 0,

            "nombre_photos": 0,

            "nombre_photos_proprietaire": 0,

            "lundi": "",
            "mardi": "",
            "mercredi": "",
            "jeudi": "",
            "vendredi": "",
            "samedi": "",
            "dimanche": ""
        }

    # ========================================================
    # COMPTAGE
    # ========================================================

    if resultat["recherche"] == "trouvé":

        nombre_trouve += 1

    elif resultat["recherche"] == "probable":

        nombre_probable += 1

    else:

        nombre_non_trouve += 1

    # ========================================================
    # AJOUT DANS EXCEL
    # ========================================================

    feuille_sortie.append([

        # Données originales
        location,
        adresse,
        code_postal,
        ville,
        pays,

        # Résultat
        resultat["recherche"],

        # Google
        resultat["adresse"],
        resultat["contact"],

        # Site web
        resultat["site_web"],

        # Date dernier post
        resultat["date_dernier_post"],

        # Avis Google
        resultat["note_google"],
        resultat["nombre_avis_google"],

        # Photos
        resultat["nombre_photos"],
        resultat["nombre_photos_proprietaire"],

        # Horaires
        resultat["lundi"],
        resultat["mardi"],
        resultat["mercredi"],
        resultat["jeudi"],
        resultat["vendredi"],
        resultat["samedi"],
        resultat["dimanche"],

        # Score
        resultat["score"]
    ])

    # ========================================================
    # ALIGNEMENT
    # ========================================================

    derniere_ligne = (
        feuille_sortie.max_row
    )

    for cellule in feuille_sortie[
        derniere_ligne
    ]:

        cellule.alignment = Alignment(

            vertical="top",

            wrap_text=True
        )

    # ========================================================
    # SCORE EN POURCENTAGE
    # ========================================================

    feuille_sortie.cell(

        derniere_ligne,

        COLONNE_SCORE

    ).number_format = "0%"


# ============================================================
# 19. LARGEUR DES COLONNES
# ============================================================

largeurs = {

    "A": 30,
    "B": 35,
    "C": 15,
    "D": 30,
    "E": 15,

    "F": 18,

    "G": 40,
    "H": 20,

    "I": 40,

    "J": 25,

    "K": 15,
    "L": 20,

    "M": 20,
    "N": 28,

    "O": 22,
    "P": 22,
    "Q": 22,
    "R": 22,
    "S": 22,
    "T": 22,
    "U": 22,

    "V": 12
}


for colonne, largeur in largeurs.items():

    feuille_sortie.column_dimensions[
        colonne
    ].width = largeur


# ============================================================
# 20. HAUTEUR DES LIGNES
# ============================================================

for numero_ligne in range(

    2,

    feuille_sortie.max_row + 1

):

    feuille_sortie.row_dimensions[
        numero_ligne
    ].height = 40


# ============================================================
# 21. SAUVEGARDE
# ============================================================

try:

    classeur_sortie.save(
        FICHIER_SORTIE
    )

except PermissionError:

    print("\n")

    print(
        "🔴 ERREUR : impossible de sauvegarder le fichier."
    )

    print(
        f"Le fichier '{FICHIER_SORTIE}' est probablement "
        "ouvert dans Excel."
    )

    print(
        "➡️ Ferme le fichier Excel puis relance le script."
    )

    raise


# ============================================================
# 22. RÉSUMÉ
# ============================================================

print("\n")

print(
    "=" * 60
)

print(
    "TRAITEMENT TERMINÉ"
)

print(
    "=" * 60
)

print(
    f"Nombre de lignes analysées : "
    f"{nombre_total}"
)

print(
    f"🟢 Trouvé (> 50 %)          : "
    f"{nombre_trouve}"
)

print(
    f"🟡 Probable (0–50 %)        : "
    f"{nombre_probable}"
)

print(
    f"🔴 Non trouvé (0 %)         : "
    f"{nombre_non_trouve}"
)

print(
    f"\nFichier créé : "
    f"{FICHIER_SORTIE}"
)

print(
    "=" * 60
)